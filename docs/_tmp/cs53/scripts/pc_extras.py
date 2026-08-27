import openpyxl, os, glob
src = r"C:\Projects\GRO.WTH\docs\_tmp\cs53\When all bonds are broken"
for fn in sorted(glob.glob(os.path.join(src, "*.xlsx"))):
    name = os.path.basename(fn)
    try:
        wb = openpyxl.load_workbook(fn, data_only=True)
        ws = wb["Main"]
        print(f"\n--- {name} ---")
        # Goals / Vines (rows 19-21 left side), KV ladder maybe later
        for r in range(17,40):
            for c_idx in range(2,14):
                v = ws.cell(row=r,column=c_idx).value
                if v is None: continue
                s = str(v).strip()
                if s in ("False","0","0.0","#N/A","None","#REF!"): continue
                if len(s) < 4 and s.isdigit(): continue
                # only print stringy content
                if any(ch.isalpha() for ch in s) and len(s) > 3:
                    coord = ws.cell(row=r,column=c_idx).coordinate
                    print(f"  {coord}: {s[:200]}")
    except Exception as e:
        print(f"ERR {name}: {e}")
